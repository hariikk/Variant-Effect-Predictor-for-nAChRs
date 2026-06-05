"""
Export all found papers to an Excel sheet with clickable DOI links.
Usage: python export_paper_list.py
"""

import pandas as pd
from openpyxl.utils import get_column_letter

import config
from checkpoint import CheckpointManager
from pubmed_search import search_all_subunits


def main():
    checkpoint = CheckpointManager()

    print("Searching PubMed for all subunits...")
    papers = search_all_subunits(config.ALL_SUBUNITS, checkpoint=checkpoint)

    if not papers:
        print("No papers found.")
        return

    # Build rows
    rows = []
    for pmid, paper in papers.items():
        doi_url = f"https://doi.org/{paper.doi}" if paper.doi else ""
        pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
        rows.append({
            "PMID": pmid,
            "Title": paper.title,
            "Authors": paper.authors,
            "Subunits": ", ".join(paper.associated_subunits),
            "DOI": paper.doi,
            "DOI Link": doi_url,
            "PubMed Link": pubmed_url,
            "PMCID": paper.pmcid,
            "Abstract": paper.abstract[:500] if paper.abstract else "",
        })

    df = pd.DataFrame(rows)
    df.sort_values(by=["Subunits", "PMID"], inplace=True)

    output_path = config.OUTPUT_DIR / "all_papers_list.xlsx"

    # Write with openpyxl so we can make links clickable
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Papers")
        ws = writer.sheets["Papers"]

        # Find column indices (1-based)
        headers = list(df.columns)
        doi_link_col = headers.index("DOI Link") + 1
        pubmed_link_col = headers.index("PubMed Link") + 1

        # Make DOI and PubMed links clickable
        for row_idx in range(2, len(df) + 2):  # row 1 is header
            # DOI link
            doi_cell = ws.cell(row=row_idx, column=doi_link_col)
            if doi_cell.value:
                doi_cell.hyperlink = doi_cell.value
                doi_cell.style = "Hyperlink"

            # PubMed link
            pm_cell = ws.cell(row=row_idx, column=pubmed_link_col)
            if pm_cell.value:
                pm_cell.hyperlink = pm_cell.value
                pm_cell.style = "Hyperlink"

        # Auto-width for key columns (not abstract)
        for col_idx, col_name in enumerate(headers, 1):
            if col_name == "Abstract":
                ws.column_dimensions[get_column_letter(col_idx)].width = 50
            elif col_name == "Title":
                ws.column_dimensions[get_column_letter(col_idx)].width = 60
            elif col_name == "Authors":
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif col_name in ("DOI Link", "PubMed Link"):
                ws.column_dimensions[get_column_letter(col_idx)].width = 45
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

    print(f"\nExported {len(df)} papers to: {output_path}")


if __name__ == "__main__":
    main()
