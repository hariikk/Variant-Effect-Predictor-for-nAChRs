"""
Extract mutation data from the shortlisted worklist papers using the recovered
fetch + Gemini pipeline, writing rows in the nachr_db_manual.xlsx format into this
human_automation/ folder. Paywalled / inaccessible papers are logged and skipped.

Usage:
    python extract_main.py --limit 5      # small test batch
    python extract_main.py                # full run (resumable)
    python extract_main.py --resume       # continue after interruption
    python extract_main.py --reset        # clear the extraction checkpoint
    python extract_main.py --no-skip-done # also process papers already in manual DB
"""

import argparse
import sys

from openpyxl import load_workbook
from tqdm import tqdm

import config
from checkpoint import CheckpointManager
from excel_writer import ExcelWriter
from llm_extractor import extract_mutations_from_paper, setup_gemini
from paper_fetcher import PaperFetcher


def load_worklist() -> list[dict]:
    """Read the `Worklist` sheet into a list of paper dicts (already score-sorted)."""
    wb = load_workbook(config.WORKLIST_EXCEL, read_only=True, data_only=True)
    ws = wb["Worklist"]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    papers = []
    for r in rows:
        papers.append({
            "pmid": str(cell(r, "PMID") or "").strip(),
            "pmcid": str(cell(r, "PMCID") or "").strip(),
            "doi": str(cell(r, "DOI") or "").strip(),
            "title": str(cell(r, "Title") or "").strip(),
            "authors": str(cell(r, "Authors") or "").strip(),
            "abstract": str(cell(r, "Abstract") or "").strip(),
            "subunits": str(cell(r, "Subunits") or "").strip(),
            "status": str(cell(r, "Status") or "").strip(),
        })
    wb.close()
    return papers


def paper_key(p: dict) -> str:
    """Stable id for checkpointing (PMID, else DOI, else title)."""
    return p["pmid"] or (f"DOI:{p['doi']}" if p["doi"] else f"TITLE:{p['title'][:60]}")


def parse_args():
    ap = argparse.ArgumentParser(
        description="Extract mutation rows from the worklist papers via Gemini.")
    ap.add_argument("--limit", type=int, default=None,
                    help="Process at most N papers (test batch).")
    ap.add_argument("--resume", action="store_true",
                    help="Continue from the extraction checkpoint (default behaviour anyway).")
    ap.add_argument("--reset", action="store_true",
                    help="Clear the extraction checkpoint and exit.")
    ap.add_argument("--no-skip-done", action="store_true",
                    help="Also process papers already in the manual DB.")
    return ap.parse_args()


def main():
    args = parse_args()
    checkpoint = CheckpointManager(config.EXTRACT_CHECKPOINT)

    if args.reset:
        checkpoint.reset()
        print("Extraction checkpoint cleared.")
        return

    if not config.NCBI_EMAIL:
        print("NCBI_EMAIL is not set in .env.")
        sys.exit(1)
    if config.EXTRACT_BACKEND == "gemini" and not config.GEMINI_API_KEY:
        print("GEMINI_API_KEY is not set. Add it to .env to run extraction.")
        sys.exit(1)
    if not config.WORKLIST_EXCEL.exists():
        print(f"Worklist not found: {config.WORKLIST_EXCEL}\nRun `python main.py --all` first.")
        sys.exit(1)

    setup_gemini()
    fetcher = PaperFetcher()
    writer = ExcelWriter()

    papers = load_worklist()
    todo = []
    for p in papers:
        if checkpoint.is_processed(paper_key(p)):
            continue
        if not args.no_skip_done and p["status"] == "already in manual DB":
            continue
        todo.append(p)
    if args.limit:
        todo = todo[: args.limit]

    print(f"Worklist: {len(papers)} papers")
    print(f"To process this run: {len(todo)} "
          f"(skipping already-done{'' if args.no_skip_done else ' + already-in-manual-DB'})")
    print(f"Output: {config.EXTRACTED_EXCEL}\n")

    processed = inaccessible = total_rows = errors = 0
    for p in tqdm(todo, desc="Extracting"):
        key = paper_key(p)
        try:
            text, source = fetcher.fetch_full_text(p["pmid"], p["pmcid"], p["doi"])
            if text is None:
                writer.append_inaccessible(
                    pmid=p["pmid"], doi=p["doi"], title=p["title"], authors=p["authors"],
                    abstract=p["abstract"],
                    subunits=[s.strip() for s in p["subunits"].split(",") if s.strip()],
                    reason=source,
                )
                checkpoint.mark_inaccessible(key, source)
                inaccessible += 1
                continue

            mutations = extract_mutations_from_paper(text, p["pmid"] or key)
            if mutations:
                doi_url = f"https://doi.org/{p['doi']}" if p["doi"] else ""
                writer.append_mutations(mutations, p["pmid"], doi_url)
                total_rows += len(mutations)
            checkpoint.mark_processed(key, len(mutations))
            processed += 1

        except KeyboardInterrupt:
            print("\nInterrupted — progress saved. Re-run to resume.")
            break
        except Exception as e:
            print(f"\n  Error on {key}: {e}")
            checkpoint.mark_error(key, str(e))
            errors += 1

    print("\n" + "=" * 60 + "\nSUMMARY\n" + "=" * 60)
    print(f"Papers processed:      {processed}")
    print(f"Papers inaccessible:   {inaccessible}")
    print(f"Errors:                {errors}")
    print(f"Mutation rows written: {total_rows}")
    print(f"\nExtracted data: {config.EXTRACTED_EXCEL}")
    print(f"Inaccessible:   {config.EXTRACT_INACCESSIBLE}")


if __name__ == "__main__":
    main()
