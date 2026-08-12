"""
Write the ranked paper worklist to a two-sheet Excel workbook.

Sheet "Worklist": papers scoring >= min_score, best first — the list to curate.
Sheet "Rejected": everything below the cutoff (kept so nothing is ever lost).

A "Status" column cross-references the existing human DB and the earlier paper
list by PMID, so papers you've already mined or listed are flagged.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config

_LINK_FONT = Font(color="0563C1", underline="single")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_LINK_COLS = {"DOI Link", "PubMed Link"}
_MAX_CELL = 32000  # stay under Excel's 32767-char cell limit

# Soft column widths (chars) keyed by header name.
_WIDTHS = {
    "PMID": 11, "Title": 60, "Authors": 26, "Year": 7, "Journal": 24,
    "Subunits": 16, "Sources": 18, "Score": 7, "Status": 20,
    "Mutation hits": 22, "Curated mutations": 22, "Curated effects": 50,
    "Effect signals": 28, "Ephys signals": 28, "DOI": 26, "DOI Link": 34,
    "PubMed Link": 40, "PMCID": 12, "Abstract": 90,
}


def _normalize_pmid(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value))
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def _load_pmid_set(path, column_name) -> set:
    """Read a set of PMIDs from a column of an existing .xlsx (best-effort)."""
    if not path or not path.exists():
        return set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set()
    pmids = set()
    try:
        rows = wb.worksheets[0].iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return set()
        try:
            col = list(header).index(column_name)
        except ValueError:
            return set()
        for row in rows:
            if col < len(row):
                p = _normalize_pmid(row[col])
                if p:
                    pmids.add(p)
    finally:
        wb.close()
    return pmids


def _status_for(pmid, mined, listed) -> str:
    if pmid in mined:
        return "already in manual DB"
    if pmid in listed:
        return "already listed"
    return "new"


def _row_for(paper, result, status) -> dict:
    doi = paper.doi or ""
    return {
        "PMID": paper.pmid,
        "Title": paper.title,
        "Authors": paper.authors,
        "Year": paper.year,
        "Journal": paper.journal,
        "Subunits": ", ".join(paper.associated_subunits),
        "Sources": ", ".join(getattr(paper, "sources", []) or []),
        "Score": result["score"],
        "Status": status,
        "Mutation hits": ", ".join(result["mutation_hits"]),
        "Curated mutations": ", ".join(getattr(paper, "curated_mutations", []) or []),
        "Curated effects": (" | ".join(getattr(paper, "curated_effects", []) or []))[:_MAX_CELL],
        "Effect signals": ", ".join(result["effect"]),
        "Ephys signals": ", ".join(result["ephys"]),
        "DOI": doi,
        "DOI Link": f"https://doi.org/{doi}" if doi else "",
        "PubMed Link": f"https://pubmed.ncbi.nlm.nih.gov/{paper.pmid}/",
        "PMCID": paper.pmcid,
        "Abstract": (paper.abstract or "")[:_MAX_CELL],
    }


def _write_sheet(ws, rows, columns):
    ws.append(columns)
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = _WIDTHS.get(name, 18)
    ws.freeze_panes = "A2"

    for r, rowdict in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            value = rowdict.get(name, "")
            cell = ws.cell(row=r, column=c, value=value)
            if name in _LINK_COLS and value:
                cell.hyperlink = value
                cell.font = _LINK_FONT


def write_worklist(scored, min_score, path=config.WORKLIST_EXCEL):
    """
    scored: iterable of (PaperMetadata, score_result dict), any order.

    Splits into Worklist (score >= min_score) and Rejected (< min_score), each
    sorted by score desc then PMID asc, and writes the workbook to `path`.
    Returns (n_worklist, n_rejected).
    """
    mined = _load_pmid_set(config.MANUAL_DB, "Reference(PMID)")
    listed = _load_pmid_set(config.PRIOR_WORKLIST, "PMID")

    rows_keep, rows_drop = [], []
    for paper, result in scored:
        status = _status_for(str(paper.pmid), mined, listed)
        row = _row_for(paper, result, status)
        (rows_keep if result["score"] >= min_score else rows_drop).append(row)

    def sort_key(row):
        pmid = str(row["PMID"])
        return (-row["Score"], int(pmid) if pmid.isdigit() else 0)

    rows_keep.sort(key=sort_key)
    rows_drop.sort(key=sort_key)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Worklist"
    _write_sheet(ws1, rows_keep, config.WORKLIST_COLUMNS)
    _write_sheet(wb.create_sheet("Rejected"), rows_drop, config.WORKLIST_COLUMNS)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return len(rows_keep), len(rows_drop)
