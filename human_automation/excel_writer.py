"""
Excel output handler for writing extracted mutation data and inaccessible paper metadata.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

import config

_LINK_FONT = Font(color="0563C1", underline="single")


def _make_doi_clickable(path):
    """Turn the bare DOI cells in an .xlsx into clickable https://doi.org links."""
    try:
        wb = load_workbook(path)
    except Exception:
        return
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "DOI" not in header:
        return
    dcol = header.index("DOI") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=dcol)
        doi = str(cell.value or "").strip()
        if not doi:
            continue
        cell.hyperlink = doi if doi.lower().startswith("http") else f"https://doi.org/{doi}"
        cell.font = _LINK_FONT
    wb.save(path)


class ExcelWriter:
    """Handles reading/writing the output Excel files."""

    def __init__(
        self,
        main_path=config.EXTRACTED_EXCEL,
        inaccessible_path=config.EXTRACT_INACCESSIBLE,
    ):
        self.main_path = main_path
        self.inaccessible_path = inaccessible_path
        self._next_oid = self._get_next_oid()

    def _get_next_oid(self) -> int:
        """Determine the next OID by reading existing data."""
        # Check the main scraper output
        if self.main_path.exists():
            try:
                df = pd.read_excel(self.main_path)
                if not df.empty and "OID" in df.columns:
                    return int(df["OID"].max()) + 1
            except Exception:
                pass

        # Also check the original manual database for max OID
        manual_db = config.MANUAL_DB
        if manual_db.exists():
            try:
                df = pd.read_excel(manual_db)
                if not df.empty and "OID" in df.columns:
                    return int(df["OID"].max()) + 1
            except Exception:
                pass

        return 1

    def append_mutations(self, mutations: list[dict], pmid: str, doi: str):
        """
        Append extracted mutations to the main Excel file.

        Args:
            mutations: List of validated mutation dicts from llm_extractor.
            pmid: PubMed ID of the source paper.
            doi: DOI of the source paper.
        """
        if not mutations:
            return

        # Build rows
        rows = []
        for mut in mutations:
            row = {
                "OID": self._next_oid,
                "nAChR subunit": mut["subunit"],
                "Modification type": mut["modification_type"],
                "AA position": mut["aa_position"],
                "Initial AA": mut["initial_aa"],
                "New AA": mut["new_aa"],
                "Effect": mut["effect"],
                "Measuring Technique": mut["measuring_technique"],
                "Pathology": mut["pathology"],
                "Reference(PMID)": pmid,
                "Entry by": config.DEFAULT_ENTRY_BY,
                "Correct?": config.DEFAULT_CORRECT,
                "DOI": doi if doi else "",
            }
            rows.append(row)
            self._next_oid += 1

        new_df = pd.DataFrame(rows, columns=config.EXCEL_COLUMNS)

        # Append to existing file or create new one
        if self.main_path.exists():
            try:
                existing_df = pd.read_excel(self.main_path)
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            except Exception:
                combined_df = new_df
        else:
            combined_df = new_df

        combined_df.to_excel(self.main_path, index=False)

    def append_inaccessible(
        self,
        pmid: str,
        doi: str,
        title: str,
        authors: str,
        abstract: str,
        subunits: list[str],
        reason: str,
    ):
        """Add a paper to the inaccessible papers Excel file."""
        row = {
            "PMID": pmid,
            "DOI": doi,
            "Title": title,
            "Authors": authors,
            "Abstract": abstract,
            "Associated Subunits": ", ".join(subunits),
            "Reason": reason,
        }

        new_df = pd.DataFrame([row], columns=config.INACCESSIBLE_COLUMNS)

        if self.inaccessible_path.exists():
            try:
                existing_df = pd.read_excel(self.inaccessible_path)
                # Avoid duplicate entries
                if pmid in existing_df["PMID"].astype(str).values:
                    return
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            except Exception:
                combined_df = new_df
        else:
            combined_df = new_df

        combined_df.to_excel(self.inaccessible_path, index=False)
        _make_doi_clickable(self.inaccessible_path)

    def get_stats(self) -> dict:
        """Return statistics about the output files."""
        stats = {"main_entries": 0, "inaccessible_papers": 0}

        if self.main_path.exists():
            try:
                df = pd.read_excel(self.main_path)
                stats["main_entries"] = len(df)
            except Exception:
                pass

        if self.inaccessible_path.exists():
            try:
                df = pd.read_excel(self.inaccessible_path)
                stats["inaccessible_papers"] = len(df)
            except Exception:
                pass

        return stats
